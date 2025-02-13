import openpyxl
from datatransfer import Data_transfer
from utilities import writing_to_file
import config


class s3_cost_assisments(Data_transfer):
    request_cost = 0.0034 # this price is for 10000 requests
    IA_retrival_cost = 0.01 # infrequent access retrival pricing GB for month

    standard_storage_cost = 0.0255 #gb per month
    infrequent_access_cost = 0.01 #gb per month
    archive_storage_cost = 0.0026 #gb per month

    def __init__(self,workbook):
        self.workbook = workbook

    def request_calculation(self,counter):
        # pricing_unit = True if self.workbook.cell(row=counter,column=15).value == "Requests" else False
        usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
        total_cost =  (usage_amount/10000) * self.request_cost
        # if pricing_unit:
        writing_to_file(self.workbook,
                            counter,
                                self.request_cost,
                                total_cost,
                                "Object storage requests $0.0034 for 10000 requests",
                                "requests",
                                "B91627")
            
    def retrival_calculation(self,counter,usagetype_tolist):

        if 'gir' in usagetype_tolist:
            writing_to_file(self.workbook,
                            counter,
                            0,
                            0,
                            "$0.00 cost for retrival, you billed for Standard tier as restored objects reside in standard tier",
                            "object storage retrival")

        elif 'zia' or 'sia' in usagetype_tolist:
            usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
            total_cost = (usage_amount*self.IA_retrival_cost)
            writing_to_file(self.workbook,
                            counter,
                            self.IA_retrival_cost,
                            total_cost,
                            "$0.01 per gb retrieved per month from infrequent access",
                            "object storage retrival",
                            "B93001")
        
    def storage_calculation(self,counter,usagetype_to_list):
        # print(f" row {counter}  and {usagetype_to_list}")
        if ('gda' in  usagetype_to_list and 'bytehrs' in usagetype_to_list):
            usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
            total_cost = usage_amount * self.archive_storage_cost
            writing_to_file(self.workbook,
                                counter,
                                 self.archive_storage_cost,
                                 total_cost,
                                 "$0.0026 per GB per month in archival storage",
                                 "glacier deeparchival storage - Archive",
                                 "B91633")
        
        if ('gir' in usagetype_to_list and ('bytehrs' in usagetype_to_list or 'smobjects' in usagetype_to_list)) :
            usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
            total_cost = usage_amount * self.archive_storage_cost
            writing_to_file(self.workbook,
                                counter,
                                 self.archive_storage_cost,
                                 total_cost,
                                 "$0.0026 per GB per month in archival storage",
                                 "glacier instant retrival storage - Archive",
                                 "B91633")
        
        # if 'glacierbytehrs' or 'glacierstaging' in usagetype_to_list  and not 'gda' or 'xy' or 'gir' in usagetype_to_list:
        if ('glacierbytehrs' in usagetype_to_list or 'glacierstaging' in usagetype_to_list) and not ('gda' in usagetype_to_list or 'xy' in usagetype_to_list or 'gir' in usagetype_to_list):
            usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
            total_cost = usage_amount * self.archive_storage_cost
            writing_to_file(self.workbook,
                                counter,
                                 self.archive_storage_cost,
                                 total_cost,
                                 "$0.0026 per GB per month in archival storage",
                                 "glacier flexible retrival storage- Archive",
                                 "B91633")
            
        #  intellegent tiering 
        if 'int' in usagetype_to_list:
            usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
            total_cost = usage_amount * self.infrequent_access_cost
            writing_to_file(self.workbook,
                                counter,
                                 self.infrequent_access_cost,
                                 total_cost,
                                 "$0.01 per GB per month in infrequent access storage",
                                 "intellegent tiering storage - infrequent",
                                 "B93000")

        # Reduced redundancy storage mapped to standard storage            
        if 'rrs' in usagetype_to_list:
            usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
            total_cost = usage_amount * self.standard_storage_cost
            writing_to_file(self.workbook,
                                counter,
                                 self.standard_storage_cost,
                                 total_cost,
                                 "$0.0255 per GB per month in standard storage",
                                 "reduced redundacy storage - Standard",
                                 "B91628")
        
        # infrequent access mapped to infrequent access
        if ('sia' in usagetype_to_list and ('bytehrs' in usagetype_to_list or 'smobjects' in usagetype_to_list)):
            usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
            total_cost = usage_amount * self.infrequent_access_cost
            writing_to_file(self.workbook,
                                counter,
                                 self.infrequent_access_cost,
                                 total_cost,
                                 "$0.1 per GB per month in infrequent storage",
                                 "infrequent access - infrequent",
                                 "B93000")
            
        # One zone -IA storage maped to infrequent access in OCI
        if ('zia' in usagetype_to_list and ('bytehrs' in usagetype_to_list or 'smobjects' in usagetype_to_list)):
            usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
            total_cost = usage_amount * self.infrequent_access_cost
            writing_to_file(self.workbook,
                                counter,
                                 self.infrequent_access_cost,
                                 total_cost,
                                 "$0.01 per GB per month in infrequent storage",
                                 "One zone infrequent access - infrequent",
                                 "B93000")
        
        # one zone express storage mapped to standard storage
        if 'xz' in usagetype_to_list and 'bytehrs' in usagetype_to_list:
            usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
            total_cost = usage_amount * self.standard_storage_cost
            writing_to_file(self.workbook,
                                counter,
                                 self.standard_storage_cost,
                                 total_cost,
                                 "$0.0255 per GB per month in standard storage",
                                 "one zone express storage - Standard",
                                 "B91628")
        
        # standard storage
        if usagetype_to_list[1] == 'timedstorage' and usagetype_to_list[2] == 'bytehrs':
            usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
            total_cost = usage_amount * self.standard_storage_cost
            writing_to_file(self.workbook,
                                counter,
                                 self.standard_storage_cost,
                                 total_cost,
                                 "$0.0255 per GB per month in standard storage",
                                 "standard storage - standard",
                                 "B91628")

        # standard storage
        if usagetype_to_list[0] == 'timedstorage' and usagetype_to_list[1] == 'bytehrs':
            usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
            total_cost = usage_amount * self.standard_storage_cost
            writing_to_file(self.workbook,
                            counter,
                                 self.standard_storage_cost,
                                 total_cost,
                                 "$0.0255 per GB per month in standard storage",
                                 "standard storage - standard",
                                 "B91628")
        
    def tagStorage_calculation(self,counter,usagetype_tolist):
        writing_to_file(self.workbook,counter,0,0,"Object storage tags","tags")
    
    def storageMonitoring_calculation(self,counter):
        writing_to_file(self.workbook,counter,0,0," OCI object storage Autotiering","Autotiering free in OCI - storage monitoring")
    
    def storageInventory_calculation(self,counter):
        writing_to_file(self.workbook,counter,0,0,"OCI object storage","NO pricing for Number of objects listed - storage inventory ")

    def storageLens_calculation(self,counter):
        writing_to_file(self.workbook,counter,0,0,"OCI object storage Metrics"," NO pricing for metrics  - storageLens")
    


    def earlyDelete_calculation(self,counter,usagetype_tolist):

        if "earlydelete" in usagetype_tolist and 'sia' in usagetype_tolist:
            usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
            total_cost = usage_amount * self.infrequent_access_cost
            writing_to_file(self.workbook,
                    counter,
                    self.infrequent_access_cost,
                    total_cost,
                    "$0.1 per GB per month in infrequent storage",
                    "Earlydelete from SIA - full storage cost charged for 31 days",
                    "B93000")
            


        elif "earlydelete" in usagetype_tolist and ('gir' in usagetype_tolist or 'bytehrs' in usagetype_tolist):
            usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
            total_cost = usage_amount * self.archive_storage_cost
            writing_to_file(self.workbook,
                    counter,
                    self.archive_storage_cost,
                    total_cost,
                    "$0.0026 per GB per month in archival storage",
                    "Earlydelete from GIR - full storage cost charged for 90 days",
                    "B91633")
    

    def process(self):
        counter = 0
        for row in range(self.workbook.max_row+1):
            counter+=1
            cell_obj = self.workbook.cell(row=counter,column=config.line_item_usage_type_column)
            if cell_obj.value != None:
                usagetype = cell_obj.value
                usagetype_tolist = usagetype.lower().split('-')

                if 'requests' in usagetype_tolist:
                    self.request_calculation(counter)
                
                if 'retrieval' in usagetype_tolist:
                    self.retrival_calculation(counter,usagetype_tolist)

                if 'timedstorage' in usagetype_tolist:
                    self.storage_calculation(counter,usagetype_tolist)
                
                if 'tagstorage' in usagetype_tolist and 'taghrs' in usagetype_tolist:
                    self.tagStorage_calculation(counter,usagetype_tolist)
                
                if 'monitoring' in usagetype_tolist and 'automation' in usagetype_tolist:
                    self.storageMonitoring_calculation(counter)

                if "inventory" in usagetype_tolist:
                    self.storageInventory_calculation(counter)
                
                if "storagelens" in usagetype_tolist or "storagelensfreetier" in usagetype_tolist and "objcount" in usagetype_tolist:
                    self.storageLens_calculation(counter)
                
                if "earlydelete" in usagetype_tolist:
                    self.earlyDelete_calculation(counter,usagetype_tolist)

        print(counter)
        
        
        