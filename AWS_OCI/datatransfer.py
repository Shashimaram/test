import openpyxl
from utilities import writing_to_file
import re
import config

# Assumptions:
# Outbound Data Transfer - Originating in North America, Europe, and UK == region1 price = $0.0085
# Outbound Data Transfer - Originating in APAC, Japan, and South America  == region2 = $0.025
# Outbound Data Transfer - Originating in Middle East and Africa == region3 = $0.05

class Data_transfer():
    """ 
    """
    region_mapping = {
        "us" : "region1",
        "na" : "region1",
        "eu" : "region1",
        "ca" : "region1",
        "da" : "region1",
        "ap" : "region2",
        "jp" : "region2",
        "sa" : "region2",
        "au" : "region2",
        "af" : "region3", 
        "il" : "region3",
        "me" : "region3"
    }

    comment = {
            "region1": "Outbound Data Transfer - Originating in North America, Europe, and UK",
            "region2": "Outbound Data Transfer - Originating in APAC, Japan, and South America",
            "region3": "Outbound Data Transfer - Originating in Middle East and Africa"
    }


    # region1 = 0.0085
    # region2 = 0.025
    # region3 = 0.05

    def __init__(self,workbook):
        self.workbook = workbook

    def calculate_cost(self, data_transfered, region):
        cost_per_gb = {
            "region1": 0.0085,
            "region2": 0.025,
            "region3": 0.05
        }
        total_cost = cost_per_gb[region] * data_transfered
        return cost_per_gb[region], total_cost
    
    def inbound_data(self,row):
        writing_to_file(workbook=self.workbook,
                            row=row,
                             column_10_value=0,
                             column_11_value=0,
                             column_12_value="inbound",
                             column_13_value="data transfer"
                             )

    def regional_data(self,row):
        writing_to_file(workbook=self.workbook,
                        row=row,
                        column_10_value=0,
                        column_11_value=0,
                        column_12_value="Regional transfer in/out",
                        column_13_value="data transfer")
    
    def outbound_data(self,counter,usagetype_tolist):
        x = usagetype_tolist[0]
        region = x[:2]
        dataTransfered = self.workbook.cell(row=counter,column=config.usage_amount_column).value
        try:
            regionValue = self.region_mapping[region]

            if regionValue == 'region1':
    
                costPerGB,totalcost = self.calculate_cost(dataTransfered,"region1")
                writing_to_file(workbook=self.workbook,
                                row=counter,
                                column_10_value=costPerGB,
                                column_11_value=totalcost,
                                column_12_value=self.comment["region1"],
                                column_13_value="data transfer",
                                skuid="B88327")

            elif regionValue == "region2":
            
                costPerGB,totalcost = self.calculate_cost(dataTransfered,"region2")
                writing_to_file(workbook=self.workbook,
                                row=counter,
                                column_10_value=costPerGB,
                                column_11_value=totalcost,
                                column_12_value=self.comment["region2"],
                                column_13_value="data transfer",
                                skuid="B93455")
            
            elif regionValue == "region3":
                costPerGB,totalcost = self.calculate_cost(dataTransfered,"region3")           
                writing_to_file(workbook=self.workbook,
                                row=counter,
                                column_10_value=costPerGB,
                                column_11_value=totalcost,
                                column_12_value=self.comment["region3"],
                                column_13_value="data transfer",
                                skuid="B93456")
                
        except Exception as e:
            print(f"error {e} at {counter}")

    def process(self):
        counter = 0
        for row in range(self.workbook.max_row+1):
            counter+=1
            cell_obj = self.workbook.cell(row=counter,column=config.line_item_usage_type_column)
            if cell_obj.value != None:
                usagetype = cell_obj.value
                # print(usagetype)
                usagetype_tolist = usagetype.lower().split('-')

                # if 'in' in usagetype_tolist and 'xaz' not in usagetype_tolist and 'bytes' in usagetype_tolist or 'abytes' in usagetype_tolist:
                if ('in' in usagetype_tolist and ('bytes' in usagetype_tolist or 'abytes' in usagetype_tolist)) and 'xaz' not in usagetype_tolist:
                    self.inbound_data(counter)
                
                # elif 'out' in usagetype_tolist and 'xaz' not in usagetype_tolist and 'bytes' in usagetype_tolist or 'abytes' in 
                elif ('out' in usagetype_tolist and ('bytes' in usagetype_tolist or 'abytes' in usagetype_tolist))and 'xaz' not in usagetype_tolist:
                    self.outbound_data(counter,usagetype_tolist)

                elif 'datatransfer' and 'xaz' in usagetype_tolist and 'bytes' in usagetype_tolist:
                    self.regional_data(counter)
                elif 'datatransfer' and 'regional' in usagetype_tolist and 'bytes' in usagetype_tolist:
                    self.regional_data(counter)
                
                elif "dataxfer" in usagetype_tolist and "out" in usagetype_tolist: #direct connect public cloud data transfer 
                    self.outbound_data(counter,usagetype_tolist)

                elif "dataxfer" in usagetype_tolist and "in" in usagetype_tolist: # direct connect public cloud data transfer 
                    self.inbound_data(counter)
                    
                elif "dataxfer" in usagetype_tolist:  # direct connect private cloud data transfer 
                    usagetype_split = re.split('[-:]',usagetype)
                    # print(usagetype_split,end="")
                    # print(counter)

                    if "DataXfer" in usagetype_split and "In" in usagetype_split:
                        writing_to_file(workbook=self.workbook,
                                        row=counter,
                                        column_10_value=0,
                                        column_11_value=0,
                                        column_12_value="private cloud inbound data transfer is free",
                                        column_13_value="data transfer")

                    if "DataXfer" in usagetype_split and "Out" in usagetype_split:
                        writing_to_file(workbook=self.workbook,
                                        row=counter,
                                        column_10_value=0,
                                        column_11_value=0,
                                        column_12_value="private cloud Outbound data transfer is free",
                                        column_13_value="data transfer")