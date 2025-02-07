import openpyxl
from utilities import writing_to_file
import re
import config

class Elastic_load_Balancing():
    
    def __init__(self,workbook):
        self.workbook = workbook
    
    def networkLoadbalancer(self,counter):
        writing_to_file(self.workbook,
                        counter,
                        0,
                        0,
                        "OCI flexible Network Loadbalancer",
                        "Network Load balancer Free in oci")
        
    def applicationLoadBalancer(self,counter):
        usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
        total_cost = usage_amount * config.Load_Balancer_Hour
        writing_to_file(self.workbook,
                        counter,
                        config.Load_Balancer_Hour,
                        total_cost,
                        "OCI flexible Loadbalancer",
                        "Loadbalancer at Layer 7")
    
    def dataprocessing(self,counter):
        writing_to_file(self.workbook,
                counter,
                0,
                0,
                "data processing at loadbalancer",
                "NO cost for data processing at Loadbalancer")
    
    def process(self):
        counter = 0
        for row in range(self.workbook.max_row+1):
            counter+=1
            cell_obj_usage = self.workbook.cell(row=counter,column=config.line_item_usage_type_column)
            cell_obj_operation = self.workbook.cell(row=counter,column=config.line_item_operation_column)
            if cell_obj_usage.value != None:
                usagetype = cell_obj_usage.value
                operationtype = cell_obj_operation.value
                usagetype_tolist = usagetype.lower().split('-')
                
                if 'lcuusage' in usagetype_tolist or "loadbalancerusage" in usagetype_tolist:
                    operationtypeSplit = operationtype.lower().split(':')
                    
                    if len(operationtypeSplit) > 1:
                        if operationtypeSplit[1] == "network":
                            self.networkLoadbalancer(counter=counter)
                            
                        if operationtypeSplit[1] == "application":
                            self.applicationLoadBalancer(counter=counter)
                            
                    elif len(operationtypeSplit) > 0:
                        self.applicationLoadBalancer(counter=counter)
                    
                if "dataprocessing" in usagetype_tolist:
                    self.dataprocessing(counter = counter)
                    
