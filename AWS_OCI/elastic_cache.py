from openpyxl import load_workbook

from assessment_engine.services.mondodb_wrapper import MongodbConnection
import os
import requests
import json



# Load the dataset using openpyxl
file_path = r"C:\Users\anudeeps\Downloads\python_auto\paramount_ext_1\elastic_cache.xlsx"# Replace with your actual file path
workbook = load_workbook(filename=file_path)
sheet = workbook.active

# Add headers for new columns if they do not exist
headers = [
    "skuids","instance_type", "vcpu", "ram","price_per_month"]
for i, header in enumerate(headers, start=sheet.max_column + 1):
    sheet.cell(row=1, column=i, value=header)

# Extract column indexes
line_item_usage_type_col = [cell.value for cell in sheet[1]].index("line_item_usage_type") + 1
line_item_description_col = [cell.value for cell in sheet[1]].index("line_item_line_item_description") + 1
LineItemOperation = [cell.value for cell in sheet[1]].index("LineItemOperation") + 1

exec_col = [cell.value for cell in sheet[1]].index("SUM(line_item_usage_amount)") + 1
skuids_col= sheet.max_column - 4
instance_type_col = sheet.max_column - 3
vcpu_col = sheet.max_column - 2
ram_col = sheet.max_column - 1
price_per_month_col = sheet.max_column

# Define a function to extract instance_type from line_item_usage_type
def extract_instance_type(usage_type):
    if ":" in usage_type:
        instance_part = usage_type.split(":")[1]
        # Replace 'xl' with 'xlarge' only if it's not already part of 'xlarge'
        if instance_part.endswith("xl") and not instance_part.endswith("xlarge"):
            return instance_part.replace("xl", "xlarge", 1)
        return instance_part
    return None


# Fetch vCPU and RAM details from MongoDB
def extract_vcpu(instance_type):
    if instance_type:
        query = {"instance_name": instance_type.lower()}
        mongoobj = MongodbConnection()
        result = mongoobj.get_data(database="Assessment", collection="aws_elastic_cache_instance_sizes", query=query)
        return result[0].get("vCPUs") if result else "Instance Not Found"
    return "Instance Not Found"

def extract_ram(instance_type):
    if instance_type:
        query = {"instance_name": instance_type.lower()}
        mongoobj = MongodbConnection()
        result = mongoobj.get_data(database="Assessment", collection="aws_elastic_cache_instance_sizes", query=query)
        return result[0].get("Memory") if result else "Instance Not Found"
    return "Instance Not Found"

# Add new data to rows
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
    usage_type = row[line_item_usage_type_col - 1].value
    description = row[line_item_description_col - 1].value
    line_item_operation = row[LineItemOperation - 1].value

    instance_type = extract_instance_type(usage_type)
    if "backup" in instance_type.lower():
        row[price_per_month_col - 1].value = row[exec_col - 1].value*0.0255
        continue
    if "serverless"  in line_item_operation.lower():
        row[price_per_month_col - 1].value = row[exec_col - 1].value *0.0194
        continue
    vcpu = extract_vcpu(instance_type)
    ram = extract_ram(instance_type)
    exec = row[exec_col - 1].value*12
    pricePerMonth = 0


    payload = {
  "provider": "oracle",
  "tfService": "cacheWithRedis",
  "service": "cacheWithRedis",
  "nodeCount":1,
  "memoryPerNode":ram,
  "executionHours":exec }

    response = requests.post('http://mccd-infradev.matildacloud.com:30092/cost/assesment', json=payload)

    if response.status_code == 200:
        response_data = response.json()  # Parse the JSON response into a dictionary
        response_data_1 = response_data[0]
        pricePerMonth = response_data_1.get('pricePerMonth',0)
        if ram >10:
            skuids = [response_data_1.get('serviceIdentifierLowMemory', 0),response_data_1.get('serviceIdentifierHighMemory', 0)]
        else:
            skuids = [response_data_1.get('serviceIdentifierLowMemory', 0)]
        skuids = ",".join(skuids)

    # Write new values to respective columns
    row[instance_type_col - 1].value = instance_type
    row[vcpu_col - 1].value = vcpu
    row[ram_col - 1].value = ram
    row[price_per_month_col - 1].value = pricePerMonth
    row[skuids_col - 1].value = skuids

# Save the updated Excel file
output_folder = r'C:\Users\anudeeps\Downloads\python_auto\paramount_ext_1'
output_file_name = 'elastic_cache_out.xlsx'
output_file_path = os.path.join(output_folder, output_file_name)

os.makedirs(output_folder, exist_ok=True)
workbook.save(output_file_path)
workbook.close()

print(f"Processed data saved to {output_file_path}")



#gggggg


#test
