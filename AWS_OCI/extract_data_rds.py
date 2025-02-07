from openpyxl import load_workbook
from scipy.signal import ellip
from sqlalchemy.databases import postgres

from assessment_engine.services.mondodb_wrapper import MongodbConnection
import os
import requests
import json



# Load the dataset using openpyxl
file_path = r"C:\Users\anudeeps\Downloads\python_auto\di\rds_instances_in.xlsx"  # Replace with your actual file path
workbook = load_workbook(filename=file_path)
sheet = workbook.active

# Add headers for new columns if they do not exist
headers = [
    "skuids","oci_instance_type","oci_ram","instance_type", "db_type", "deployment_type", "vcpu", "ram", "ocpu","oci_vcpu","price_per_month"]
for i, header in enumerate(headers, start=sheet.max_column + 1):
    sheet.cell(row=1, column=i, value=header)

# Extract column indexes
line_item_usage_type_col = [cell.value for cell in sheet[1]].index("line_item_usage_type") + 1
line_item_description_col = [cell.value for cell in sheet[1]].index("line_item_line_item_description") + 1
exec_col = [cell.value for cell in sheet[1]].index("SUM(line_item_usage_amount)") + 1
instance_type_col = sheet.max_column - 7
db_type_col = sheet.max_column - 6
deployment_type_col = sheet.max_column - 5
vcpu_col = sheet.max_column - 4
ram_col = sheet.max_column - 3
ocpu_col = sheet.max_column - 2
oci_ram_col = sheet.max_column - 8
oci_instance_type_col = sheet.max_column - 9
skuids_col = sheet.max_column - 10
oci_vcpu_col = sheet.max_column - 1
price_per_month_col = sheet.max_column

# Define a function to extract instance_type from line_item_usage_type
def extract_instance_type(usage_type):
    if ":" in usage_type:
        instance_part = usage_type.split(":")[1]
        # Replace 'xl' with 'xlarge' only if it's not already part of 'xlarge'
        if instance_part.endswith("xl") and not instance_part.endswith("xlarge"):
            return instance_part.replace("xl", "xlarge", 1)
        return instance_part
    return None



# Define a function to extract db_type from line_item_line_item_description
def extract_db_type(description):
    db_type = 0
    db_types = ['mysql', 'sql server', 'postgres', 'oracle']
    description_lower = description.lower()
    for db_type in db_types:
        if db_type in description_lower:
            return db_type

    return None

# Define a function to extract deployment_type from line_item_usage_type
def extract_deployment_type(usage_type):
    if "multi-az" in usage_type.lower():
        return "Multi-AZ"
    return "Single-AZ"

# Fetch vCPU and RAM details from MongoDB
def extract_vcpu(instance_type):
    if instance_type:
        query = {"instance_name": instance_type.lower()}
        mongoobj = MongodbConnection()
        result = mongoobj.get_data(database="Assessment", collection="aws_rds_instance_sizes", query=query)
        return result[0].get("vcpu") if result else "Instance Not Found"
    return "Instance Not Found"

def extract_ram(instance_type):
    if instance_type:
        query = {"instance_name": instance_type.lower()}
        mongoobj = MongodbConnection()
        result = mongoobj.get_data(database="Assessment", collection="aws_rds_instance_sizes", query=query)
        return result[0].get("ram") if result else "Instance Not Found"
    return "Instance Not Found"

# Add new data to rows
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
    usage_type = row[line_item_usage_type_col - 1].value
    row_number = row
    response_data = 0
    response_data_1 = 0
    max_row_number = sheet.max_row
    description = row[line_item_description_col - 1].value

    instance_type = extract_instance_type(usage_type)
    if "db" not in instance_type.lower():
        continue
    db_type = extract_db_type(description)
    if db_type is None:
        db_type = "mysql"
    deployment_type = extract_deployment_type(usage_type)
    vcpu = extract_vcpu(instance_type)
    ram = extract_ram(instance_type)
    ocpu = vcpu / 2 if isinstance(vcpu, (int, float)) else None
    exec = row[exec_col - 1].value*12
    pricePerMonth = 0
    skuids = 0


    if db_type == "mysql":
        payload = {
        "engine": "mysql",
          "provider": "oracle",
          "deploymentType": "mysqlHeatwave",
          "serviceName": "DatabaseMySQL",
          "tfService": "mysql",
          "service": "Database",
          "CPUMetric": "ECPU",
          "ecpu" : vcpu,
          "storage":0,
          "backupStorage":0,
          "totalShapeMemoryInGB":0,
          "executionHours":exec,
          "enableHA": "yes" if "multi" in deployment_type.lower() else "linux",
        }
        skuids = "B108030"
        # Make the API request
        response = requests.post('http://mccd-infradev.matildacloud.com:30092/cost/assesment', json=payload)

        if response.status_code == 200:
            response_data = response.json()  # Parse the JSON response into a dictionary
            response_data_1 = response_data[0]
            oci_ecpu = response_data_1.get('vcpu', 0)
            oci_instance_type = "OCI eCPU"
            ocpu = 0
            oram = oci_ecpu*8
            pricePerMonth = response_data_1.get('databaseForHeatwaveStandardNodePricePerMonth', 0)

    elif db_type == "postgres":

        payload = {
                "provider": "oracle",
                "service": "Database",
                "engine": "postgresql",
                "processor": "AMD" if "g." in instance_type.lower() else "Intel",
                "vcpu": vcpu,
                "memory": ram,
                "tfService": "postgresql",
                "serviceName": "DatabasePostgreSQL",
                "subInstanceFamily": "PostgreSQL.VM.Standard.Flex.E4" if "g." in instance_type.lower() else "PostgreSQL.VM.Standard3.Flex",
                "storagePerformanceTier": "75,000 IOPS",
                "databaseOptimizedStorage": 0,
                "nodesPerCluster": 1,
                "executionHours": exec
        }


        # Make the API request
        response = requests.post('http://mccd-infradev.matildacloud.com:30092/cost/assesment', json=payload)

        if response.status_code == 200:
            response_data = response.json()  # Parse the JSON response into a dictionary
            response_data_1 = response_data[0]
            ocpu = response_data_1.get('ocpu', 0)
            oram = response_data_1.get('memory', 0)
            skuids = [response_data_1.get('serviceIdentifierComputeOCPU', 0), response_data_1.get('serviceIdentifierComputeMemory', 0)]
            skuids = ",".join(skuids)
            oci_instance_type = response_data_1.get('subInstanceFamily', 0)
            pricePerMonth = (response_data_1.get('pricePerMonthForComputeOCPU',0) + response_data_1.get('pricePerMonthForComputeMemory',0))*3  if "multi" in deployment_type.lower() else (response_data_1.get('pricePerMonthForComputeOCPU',0) + response_data_1.get('pricePerMonthForComputeMemory',0)) # Safely get the value

    elif db_type == "sql server":
        payload = {
                "provider": "oracle",
                "operatingSystem": "linux",
                "service": "compute",
                "vcpu": vcpu,
                "ram": ram,
                "processor": "Intel",
                "pricingType": "ondemand",
                "serviceType": "virtualMachine",
                "subInstanceFamily": [
                    "VM.Standard3"
                ],
                 "executionHours":exec
            }

        # Make the API request
        response = requests.post('http://mccd-infradev.matildacloud.com:30092/cost', json=payload)

        if response.status_code == 200:
            response_data = response.json()  # Parse the JSON response into a dictionary
            #response_data_1 = response_data[0]
            ocpu = response_data.get('ocpu', 0)
            oci_instance_type = response_data.get('name', 0)
            oram = response_data.get('memory', 0)
            skuids = [response_data.get('serviceIdentifierOCPU', 0),response_data.get('serviceIdentifierMemory', 0)]
            skuids = ",".join(skuids)
            pricePerMonth = response_data.get('pricePerMonth',0)*3 if "multi" in deployment_type.lower() else response_data.get('pricePerMonth',0) # Safely get the value

    elif db_type == "oracle":
        if "byol" in description.lower():
            license = "byol"
        elif "oracle ee" in description.lower():
            license = "enterprise"
        elif "oracle se" in description.lower():
            license = "standard"
        if "g." in instance_type.lower():
            oracle_processor = "Ampere"
            sub_instance = "VM.Standard.A1"
        else: #"i." in instance_type.lower():
            oracle_processor = "Intel"
            sub_instance = "VM.Standard3"
        payload = {
                "provider": "oracle",
                "service": "database",
                "tfService": "baseDatabase",
                 "type": license,
                 "vcpu": vcpu,
                #//"ocpu": 1,
                "processor": oracle_processor,
                 "subInstanceFamily": [
                sub_instance
    ],
            "totalAvailableStorage": 0,
    "performanceLevel": "balanced",
                 "executionHours":exec
            }

        # Make the API request
        response = requests.post('http://mccd-infradev.matildacloud.com:30092/cost/assesment', json=payload)

        if response.status_code == 200:
            response_data = response.json()  # Parse the JSON response into a dictionary
            response_data_1 = response_data[0]
            ocpu = response_data_1.get('ocpu', 0)
            oram = response_data_1.get('memory', 0)
            oci_instance_type = response_data_1.get('name', 0)
            skuids = response_data_1.get('serviceIdentifierOCPU', 0)
            pricePerMonth = response_data_1.get('pricePerMonthOcpu',0)*3 if "multi" in deployment_type.lower() else response_data_1.get('pricePerMonthOcpu',0) # Safely get the value




    # Write new values to respective columns
    row[oci_instance_type_col - 1].value = oci_instance_type
    row[instance_type_col - 1].value = instance_type
    row[db_type_col - 1].value = db_type
    row[deployment_type_col - 1].value = deployment_type
    row[vcpu_col - 1].value = vcpu
    row[ram_col - 1].value = ram
    row[ocpu_col - 1].value = ocpu
    row[oci_ram_col - 1].value = oram
    row[oci_vcpu_col - 1].value = oci_ecpu if db_type == "mysql" else 0
    row[price_per_month_col - 1].value = pricePerMonth
    row[skuids_col - 1].value = skuids

# Save the updated Excel file
output_folder = r'C:\Users\anudeeps\Downloads\python_auto\di'
output_file_name = 'rds_instance_out.xlsx'
output_file_path = os.path.join(output_folder, output_file_name)

os.makedirs(output_folder, exist_ok=True)
workbook.save(output_file_path)
workbook.close()

print(f"Processed data saved to {output_file_path}")
