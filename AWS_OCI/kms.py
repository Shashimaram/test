import openpyxl
from datatransfer import Data_transfer
from utilities import writing_to_file
import config

class Key_management_service():
    
    def __init__(self,workbook):
        self.workbook = workbook
    
    def process(self):
        counter = 0
        for row in range(self.workbook.max_row+1):
            counter+=1
            cell_obj = self.workbook.cell(row=counter,column=config.line_item_usage_type_column)
            if cell_obj.value != None:
                usagetype = cell_obj.value
                usagetype_tolist = usagetype.lower().split('-')
                
                if "keys" in usagetype_tolist and "kms" in usagetype_tolist:
                    usage_amount = self.workbook.cell(row=counter, column = config.usage_amount_column).value
                    total_cost = usage_amount * config.oci_kms_per_version_cost
                    writing_to_file(
                        self.workbook,
                        counter,
                        config.oci_kms_per_version_cost,
                        total_cost,
                        "OCI Key Management",
                        "OCI Key Management - Vault - HSM protected Keys"
                        )
                
                elif "requests" in usagetype_tolist and "kms" in usagetype_tolist:
                    writing_to_file(
                        self.workbook,
                        counter,
                        0,
                        0,
                        "OCI Key Management",
                        "Requests for KMS are free"
                        )