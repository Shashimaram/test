import openpyxl
from S3costcalculatins import s3_cost_assisments
from datatransfer import Data_transfer
from awsQcostcalculation import Sqs_cost_assesments
from vpcCostCalculation import VPC_cost_assessment
from reportGeneration import ReportGenerating
from lambdaCostcalculation import Lambda_cost_assessments
from directConnect import DirectConnect_cost_assessment
from elasticLoadBalancing import Elastic_load_Balancing
from kms import Key_management_service
from elementalMediaLive import ElementalMediaLive
from ec2Others import ec2_Others_cost_assessments

# print("test")

path =r"C:\Users\smaram\Downloads/extracted_data_multiple_sheets.xlsx"

ws = openpyxl.load_workbook(path)

# workbook = ws.active

# direcconnect = DirectConnect_cost_assessment(workbook=workbook)
# direcconnect.process()

print(ws.sheetnames)

for sheet in ws.sheetnames:
    paramountdataTranfer = Data_transfer(ws[sheet]) #* AWS direct connect is included as data transfer
    paramountdataTranfer.process()
    paramountEC2 = ec2_Others_cost_assessments(ws[sheet])
    paramountEC2.process()
    break
    # paramountMedia = ElementalMediaLive(ws[sheet])
    # paramountMedia.process()
    
    # elb = Elastic_load_Balancing(ws[sheet])
    # elb.process()
    
    # kms = Key_management_service(ws[sheet])
    # kms.process()
    
    

    # if sheet  == 'AmazonVPC':
    #     try:
    #         paramountVPC = VPC_cost_assessment(ws[sheet])
    #         paramountVPC.process()
    #     except Exception as e:
    #         print(f"error {e} at {sheet}")

    # if sheet == 'AmazonS3':
    #     try:
    #         paramountS3 = s3_cost_assisments(ws[sheet])
    #         paramountS3.process()
    #     except Exception as e:
    #         print(f"error {e} at {sheet}")
            
            
    # if sheet == 'ELB':
    #     try:
    #         paramountS3 = Elastic_load_Balancing(ws[sheet])
    #         paramountS3.process()
    #     except Exception as e:
    #         print(f"error {e} at {sheet}")
            
            
    # if sheet  == 'AWSDirectConnect':
    #     # continue
    #     try:
    #         direcconnect = DirectConnect_cost_assessment(ws[sheet])
    #         direcconnect.process()
    #     except Exception as e:
    #         print(f"error {e} at {sheet}")
#     if sheet == 'AWSQueueService':
#         try:
#         # continue
#             paramountQ = Sqs_cost_assesments(ws[sheet])
#             paramountQ.process()
#         except Exception as e:
#             print(f"error {e} at {sheet}")
            
    # if sheet == 'Lambda':
    #     try:
    # paramountLambda = Lambda_cost_assessments(ws[sheet])
    # paramountLambda.process()
    #         paramountdataTranfer = Data_transfer(ws[sheet]) #* AWS direct connect is included as data transfer
    #         paramountdataTranfer.process()   
    #     except Exception as e:
    #         print(f"error {e} at {sheet}")
    # pass

ws.save(path)