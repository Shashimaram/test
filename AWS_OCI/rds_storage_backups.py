from openpyxl import load_workbook

from assessment_engine.services.mondodb_wrapper import MongodbConnection
import os
import requests
import json


#from assessment_engine.services.new import price_per_month, payload

# Load the dataset using openpyxl
file_path = r"C:\Users\anudeeps\Downloads\python_auto\paramount_ext_1\rds_storage_backup.xlsx"  # Replace with your actual file path
workbook = load_workbook(filename=file_path)
sheet = workbook.active

# Add headers for new columns if they do not exist
headers = ["storage_type","skuids","category","db_type","deployment_type","unit_price","oci_cost"]
for i, header in enumerate(headers, start=sheet.max_column + 1):
    sheet.cell(row=1, column=i, value=header)

# Extract column indexes
line_item_usage_type_col = [cell.value for cell in sheet[1]].index("line_item_usage_type") + 1
line_item_description_col = [cell.value for cell in sheet[1]].index("line_item_line_item_description") + 1
exec_col = [cell.value for cell in sheet[1]].index("SUM(line_item_usage_amount)") + 1
storage_type_col = sheet.max_column - 6
skuids_col= sheet.max_column - 5
category_col = sheet.max_column - 4
db_type_col = sheet.max_column - 3
deployment_type_col = sheet.max_column - 2
unit_price_col = sheet.max_column -1
oci_cost_col = sheet.max_column

# Define a function to extract instance_type from line_item_usage_type
def extract_category_type(usage_type):
    if "backup" in usage_type.lower():
        return "backup"
    elif "storageio" in usage_type.lower():
        return "storageio"
    elif "storage" in usage_type.lower():
        return "storage"
    elif "iops" in usage_type.lower():
        return "iops"
    elif "requests" in usage_type.lower():
        return "requests"
    return None


# Fetch vCPU and RAM details from MongoDB
def extract_dbtype(description):
    db_types = ['mysql', 'sql server', 'postgres', 'oracle']
    description_lower = description.lower()
    for db_type in db_types:
        if db_type in description_lower:
            return db_type

def extract_deployment_type(description):
    if "multi-az" in description.lower():
        return "Multi-AZ"
    return "Single-AZ"


def extract_unitprice(category,db_type, deployment_type):
    if category == "storage":
        if db_type == "mysql":
            unit_price = 0.04
        elif db_type == "sql server":
            unit_price = 0.0255
        elif db_type == "postgres":
            unit_price = 0.072
        elif db_type == "oracle":
            unit_price = 0.0255
        elif db_type == "mariadb":
            unit_price = 0.04
        elif db_type == "aurora":
            unit_price = 0.04
        else:
            raise ValueError(f"Unknown database type: {db_type}")

        if deployment_type == "Multi-AZ":
            return unit_price * 3
        else:
            return unit_price
    elif category == "backup":
        if db_type == "mysql":
            unit_price = 0.04
        else:
            unit_price = 0.0255

        return  unit_price




def extract_oci_cost(category,unit_price,exec,storage_type,storage_type_value=0,):
    if storage_type == "gp2" or storage_type == "gp3":
        storage_type_value = 10
    elif storage_type == "io1" or storage_type == "io2":
        storage_type_value = 30
    elif storage_type == "st1" or storage_type == "sc1" or storage_type == "magnetic":
        storage_type_value = 1

    if "requests" in category.lower() or "iops" in category.lower():
        exec = 0.0
        unit_price = 0.0
        storage_type_value = 0.0
    return float(unit_price*exec)+float(exec*0.0017*storage_type_value)

# Add new data to rows


def extract_skuids( db_type,storage_type):
        if db_type == "mysql":
            skuids = ["B92426"]
        elif db_type == "sql server":
            skuids = ["B91961","B91962"]
        elif db_type == "postgres":
            skuids = ["B99062"]
        elif db_type == "oracle":
            skuids = ["B91961"]
        elif db_type == "mariadb":
            skuids = ["B92426"]
        elif db_type == "aurora":
            skuids = ["B92426"]
        else:
            raise ValueError(f"Unknown database type: {db_type}")
        return skuids


def extract_storage_type(description):
    if "gp2" in description.lower():
        return "gp2"
    elif "io1" in description.lower():
        return "io1"
    elif "io2" in description.lower():
        return "io2"
    elif "st1" in description.lower():
        return "st1"
    elif "sc1" in description.lower():
        return "sc1"
    elif "gp3" in description.lower():
        return "gp3"
    elif "magnetic" in description.lower():
        return "magnetic"
    else:
        return None


for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=False):
    usage_type = row[line_item_usage_type_col - 1].value
    description = row[line_item_description_col - 1].value

    category = extract_category_type(usage_type)
    if category == "requests" or category == "iops" or category == "storageio":
        row[oci_cost_col - 1].value = 0
        continue
    db_type = extract_dbtype(description)
    if db_type is None:
        db_type = "mysql"
    deployment_type = extract_deployment_type(description)
    unit_price = extract_unitprice(category,db_type,deployment_type)
    storage_type = extract_storage_type(description)
    exec = row[exec_col - 1].value
    oci_cost = extract_oci_cost(category,unit_price, exec, storage_type,storage_type_value=0)
    skuids = extract_skuids(db_type,storage_type)
    skuids = ",".join(skuids)



    # Write new values to respective columns
    row[category_col - 1].value = category
    row[db_type_col - 1].value = db_type
    row[deployment_type_col - 1].value = deployment_type
    row[unit_price_col - 1].value = unit_price
    row[oci_cost_col - 1].value = oci_cost
    row[skuids_col - 1].value = skuids
    row[storage_type_col - 1].value = storage_type


# Save the updated Excel file
output_folder = r'C:\Users\anudeeps\Downloads\python_auto\paramount_ext_1'
output_file_name = 'rds_storage_backups_out.xlsx'
output_file_path = os.path.join(output_folder, output_file_name)

os.makedirs(output_folder, exist_ok=True)
workbook.save(output_file_path)
workbook.close()

print(f"Processed data saved to {output_file_path}")
