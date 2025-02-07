import openpyxl
import os

# Function to extract rows based on list values and save them to a single file with multiple sheets
def extract_rows_to_excel(input_file, value_list,output_folder):
    # Load the input Excel file
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Get the headers (first row)
    headers = [cell.value for cell in sheet[1]]
    
    # Create a new workbook to hold all the sheets
    new_wb = openpyxl.Workbook()
    
    # Remove the default sheet created by openpyxl, we will add custom ones later
    new_wb.remove(new_wb.active)

    # Iterate over the values in the provided list
    for value in value_list:
        # Create a new sheet for each value
        new_sheet = new_wb.create_sheet(title=value)

        # Add headers to the new sheet
        for col_num, header in enumerate(headers, 1):
            new_sheet.cell(row=1, column=col_num, value=header)

        # Variable to track if any row was added
        row_added = False

        # Iterate through the rows in the input sheet (starting from row 2, to skip headers)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == value:  # Check if the first column matches the list value
                row_added = True
                new_sheet.append(row)  # Add the matching row to the new sheet

        # If no rows were added for a value, just leave that sheet empty
        if row_added:
            print(f"Added rows for: {value}")
        else:
            print(f"No matching rows for value: {value}")
        
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Define the full path to save the file
    output_file = os.path.join(output_folder, "extracted_data_multiple_sheets.xlsx")

    # Save the new workbook with multiple sheets
    # output_file = "extracted_data_multiple_sheets.xlsx"
    new_wb.save(output_file)
    print(f"Created file: {output_file}")

# Example usage
if __name__ == "__main__":
    # Input Excel file
    input_excel_file = r"C:\Users\smaram\\Downloads\Billing.xlsx"  # replace with your input file name
    output_folder = r"C:\Users\smaram\\Downloads"

    # List of values to search for in the first column
    search_list = ['AmazonS3', 'AmazonVPC', 'AWSDirectConnect', 'transcribe','AmazonEKS']  # replace with your list values

    # Call the function to extract rows and create Excel files
    extract_rows_to_excel(input_excel_file, search_list,output_folder)

