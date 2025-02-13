import openpyxl
import json

def js_r(filename: str):
    with open(filename) as f_in:
        return json.load(f_in)
    
    
path =r"C:\Users\smaram\Desktop\Book9.xlsx"
data = js_r('dataTransferSKU.json')
# print(data)

ws = openpyxl.load_workbook(path)
sheetnames = ws.sheetnames
workbook = ws[sheetnames[0]]


counter = 2
for row in range(1, workbook.max_row + 1):
    sku_id = workbook.cell(row=counter, column=5).value
    print(sku_id)
    # try:
    if data[sku_id]["costs"][0] == int:
        print(data[sku_id]["costs"][0])
        if sku_id is not None:
            unitPrice = data[sku_id]["costs"][0]
            # print(f"{sku_id} and {unitPrice}")
            if unitPrice is not None:
                workbook.cell(row=counter, column=12, value=unitPrice)
                usageAmount = workbook.cell(row=counter, column=9).value
                if usageAmount is not None:
                    totalcost = unitPrice * usageAmount
                    # print(totalcost)
                    workbook.cell(row=counter, column=13, value=totalcost)
    # except Exception as e:
    #     print(f"{e} at row {counter}")
    counter += 1
ws.save(path)               