import openpyxl

data = {
    "CF63-3CCD-F6EC": 0.023,
    "0D5D-6E23-4250": 0.026,
    "01C0-1EAA-23AD": 0.023,
    "F272-7933-F065": 0.023, 
    "EC40-8747-D6FF": 0.026,
    "E5F0-6A5D-7BAD": None,
    "A703-5CB6-E0BF": 0.02,
    "E653-0A40-3B69": 0.026,
    "7870-010B-2763": None,
    "4DBF-185F-A415": None,
    "36B5-7D3A-C3EE": 0.000001,
    "9559-D2EC-1D33": 0.000001,
    "21A9-6205-2167": 0.0005,
    "9E8A-BB82-26BF": None,
    "80A7-0A73-0413": 0.00058666,
    "1299-FB0A-5AB8": 0.0176,
    "7253-8121-9E13": 0.00058666,
    "B205-1CAE-9FB2": 0.0176,
    "7BE5-EBE7-F791": 0.0005,
    "4CF0-069A-15D9": 0.015,
    "8D92-97F6-DDF3": 0.01,
    "8AE7-5BBD-8F38": 0.0004,
    "9ADA-9AED-1B24": 0.000001,
    "5693-D6A3-C16C": 0.000001,
    "C31D-D5AC-885F": 0.00002,
    "1193-6316-413E": 0.08,
    "AED0-3315-7B11": 0.02,
    "CB83-3C2D-160D": 0.02,
    "3C8A-99C5-F47B": 0.00002,
    "D46A-868A-BBF7": 0.02,
    "990F-BF38-8D3C": 0.08,
    "3DCF-1741-642F": 0.08,
    "68F8-91DC-CFA9": 0.02,
    "07A4-C7D5-D204": None,
    "27F0-D54C-619A": 0.10,
    "C7FF-4F9E-C0DB": None,
    "6B37-399C-BF69": None,
    "1E7D-CBB0-AF0C": 0.10,
    "EF0A-B3BA-32CA": 0.08,
    "CDD1-6B91-FDF8": 0.10,
    "8DFA-C1B4-82A5": 0.000001,
    "E03D-A650-232E": 0.00002,
    "22EB-AAE8-FBCD": None,
    "9B2D-2B7D-FA5C": None,
    "1F8B-71B0-3D1B": None,
    "80E7-B02A-1CAD": 0.01,
    "2369-E523-6E72": 0.01,
    "FAFF-3200-1569": 0.04,
    "F869-7E17-8798": 0.04,
    "80E8-7294-105A": 0.04,
    "5391-66E2-9811": 0.04,
    "A6BD-E69F-4F97": 0.04,
    "85E4-9644-93A7": 0.04,
    "3712-109E-A7DC": 0.04,
    "DF96-9F56-0CF7": 0.04,
    "D980-4666-37D3": 0.04,
    "82E0-D7E0-1CB0": 0.04,
    "62CF-0844-EF78": 0.04,
    "0A5B-5BB2-9A38": 0.04,
    "12CB-A27E-C00F": 0.04,
    "3498-78F6-3EC4": 0.04
}
for x in data:
    print(data)
    
    
path =r"C:\Users\smaram\Desktop\Book9.xlsx"

ws = openpyxl.load_workbook(path)

workbook = ws.active



counter = 1
for row in range(1, workbook.max_row + 1):
    sku_id = workbook.cell(row=counter, column=5).value
    print(sku_id)
    try:
        if sku_id is not None:
            unitPrice = data.get(sku_id, None)
            print(f"{sku_id} and {unitPrice}")
            if unitPrice is not None:
                workbook.cell(row=counter, column=12, value=unitPrice)
                usageAmount = workbook.cell(row=counter, column=9).value
                if usageAmount is not None:
                    totalcost = unitPrice * usageAmount
                    print(totalcost)
                    workbook.cell(row=counter, column=13, value=totalcost)
    except Exception as e:
        print(f"{e} at row {counter}")
    counter += 1
ws.save(path)               