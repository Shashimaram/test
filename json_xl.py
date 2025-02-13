import json
from openpyxl import Workbook

def json_file_to_excel(json_file_path, output_file):
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    
    # Read JSON file
    try:
        with open(json_file_path, 'r') as file:
            json_data = json.load(file)
    except FileNotFoundError:
        print(f"Error: JSON file '{json_file_path}' not found")
        return
    except json.JSONDecodeError:
        print(f"Error: Invalid JSON format in file '{json_file_path}'")
        return
    
    # Handle both single dict and list of dicts
    if isinstance(json_data, dict):
        json_data = [json_data]
        
    # Write headers
    if json_data:
        headers = list(json_data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
    
    # Write data rows
    for row, item in enumerate(json_data, 2):
        for col, key in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=str(item.get(key, '')))
    
    # Save the workbook
    try:
        wb.save(output_file)
        print(f"Excel file successfully created at '{output_file}'")
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")

# Example usage:
"""
# Convert JSON file to Excel
json_file_to_excel('input.json', 'output.xlsx')


"""

json_file_to_excel(r'C:\Users\smaram\Desktop\azure_data.json', 'output.xlsx')

