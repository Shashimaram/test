import openpyxl
import pymongo

path =r"C:\Users\smaram\Downloads/extracted_data_multiple_sheets.xlsx"

ws = openpyxl.load_workbook(path)

# workbook = ws.active

myclient = pymongo.MongoClient("mongodb://admin:Matilda7%23@mccd-infradev.matildacloud.com:30020/")

mydb  = myclient["matildacost"]
collection = mydb["mcost_oci"]

for sheet in ws.sheetnames:
        counter = 0
        for row in range(sheet.max_row+1):
            counter+=1
            cell_obj = sheet.cell(row=counter,column=15)
            if cell_obj != None
    